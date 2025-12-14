# reportes/utils.py
"""
Utilidades para exportación de reportes
 Exportación a Excel (openpyxl)
 Exportación a CSV
 Formateo y estilos
"""
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv
from datetime import datetime
import logging

logger = logging.getLogger('reportes')


# ============================================
# EXPORTAR A EXCEL
# ============================================

def exportar_pedidos_excel(queryset):
    """
    Exporta pedidos a formato Excel con formato profesional

    Args:
        queryset: QuerySet de Pedido

    Returns:
        HttpResponse con archivo Excel
    """
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Pedidos"

    # ============================================
    # ESTILOS
    # ============================================

    # Estilo del encabezado
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Estilo de datos
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center')

    # Bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ============================================
    # ENCABEZADOS
    # ============================================
    headers = [
        'ID',
        'Tipo',
        'Estado',
        'Cliente',
        'Email Cliente',
        'Proveedor',
        'Repartidor',
        'Dirección Entrega',
        'Total',
        'Comisión Repartidor',
        'Comisión Proveedor',
        'Ganancia App',
        'Método Pago',
        'Fecha Creación',
        'Fecha Entrega',
        'Cancelado Por',
    ]

    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # ============================================
    # DATOS
    # ============================================

    # Optimizar queryset
    pedidos = queryset.select_related(
        'cliente__user',
        'proveedor',
        'repartidor__user'
    )

    for row_num, pedido in enumerate(pedidos, 2):
        # Preparar datos
        row_data = [
            pedido.id,
            pedido.get_tipo_display(),
            pedido.get_estado_display(),
            pedido.cliente.user.get_full_name(),
            pedido.cliente.user.email,
            pedido.proveedor.nombre if pedido.proveedor else 'N/A',
            pedido.repartidor.user.get_full_name() if pedido.repartidor else 'Sin asignar',
            pedido.direccion_entrega,
            float(pedido.total),
            float(pedido.comision_repartidor),
            float(pedido.comision_proveedor),
            float(pedido.ganancia_app),
            pedido.metodo_pago,
            pedido.creado_en.strftime('%Y-%m-%d %H:%M:%S'),
            pedido.fecha_entregado.strftime('%Y-%m-%d %H:%M:%S') if pedido.fecha_entregado else 'N/A',
            pedido.cancelado_por or 'N/A',
        ]

        # Escribir fila
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

            # Formato para montos
            if col_num in [9, 10, 11, 12]:  # Columnas de dinero
                cell.number_format = '$#,##0.00'

    # ============================================
    # AJUSTAR ANCHOS DE COLUMNA
    # ============================================
    column_widths = {
        'A': 8,   # ID
        'B': 15,  # Tipo
        'C': 15,  # Estado
        'D': 25,  # Cliente
        'E': 30,  # Email
        'F': 25,  # Proveedor
        'G': 25,  # Repartidor
        'H': 40,  # Dirección
        'I': 12,  # Total
        'J': 18,  # Comisión Repartidor
        'K': 18,  # Comisión Proveedor
        'L': 15,  # Ganancia App
        'M': 15,  # Método Pago
        'N': 20,  # Fecha Creación
        'O': 20,  # Fecha Entrega
        'P': 15,  # Cancelado Por
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # ============================================
    # AGREGAR FILA DE TOTALES
    # ============================================
    total_row = len(pedidos) + 2

    # Etiqueta
    ws.cell(row=total_row, column=1).value = 'TOTALES:'
    ws.cell(row=total_row, column=1).font = Font(name='Arial', size=11, bold=True)

    # Calcular totales
    totales = queryset.aggregate(
        total=csv.Sum('total'),
        comision_repartidor=csv.Sum('comision_repartidor'),
        comision_proveedor=csv.Sum('comision_proveedor'),
        ganancia_app=csv.Sum('ganancia_app')
    )

    # Total
    cell = ws.cell(row=total_row, column=9)
    cell.value = float(totales['total'] or 0)
    cell.number_format = '$#,##0.00'
    cell.font = Font(name='Arial', size=11, bold=True)
    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

    # Comisión Repartidor
    cell = ws.cell(row=total_row, column=10)
    cell.value = float(totales['comision_repartidor'] or 0)
    cell.number_format = '$#,##0.00'
    cell.font = Font(name='Arial', size=11, bold=True)
    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

    # Comisión Proveedor
    cell = ws.cell(row=total_row, column=11)
    cell.value = float(totales['comision_proveedor'] or 0)
    cell.number_format = '$#,##0.00'
    cell.font = Font(name='Arial', size=11, bold=True)
    cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

    # Ganancia App
    cell = ws.cell(row=total_row, column=12)
    cell.value = float(totales['ganancia_app'] or 0)
    cell.number_format = '$#,##0.00'
    cell.font = Font(name='Arial', size=11, bold=True)
    cell.fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')

    # ============================================
    # CONGELAR PANEL
    # ============================================
    ws.freeze_panes = 'A2'

    # ============================================
    # GENERAR RESPUESTA
    # ============================================
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    filename = f"reporte_pedidos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)

    logger.info(f"Reporte Excel generado: {len(pedidos)} pedidos")

    return response


# ============================================
# EXPORTAR A CSV
# ============================================

def exportar_pedidos_csv(queryset):
    """
    Exporta pedidos a formato CSV

    Args:
        queryset: QuerySet de Pedido

    Returns:
        HttpResponse con archivo CSV
    """
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"reporte_pedidos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Agregar BOM para Excel (soporte UTF-8)
    response.write('\ufeff')

    writer = csv.writer(response, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

    # Encabezados
    writer.writerow([
        'ID',
        'Tipo',
        'Estado',
        'Cliente',
        'Email Cliente',
        'Celular Cliente',
        'Proveedor',
        'Repartidor',
        'Dirección Entrega',
        'Total',
        'Comisión Repartidor',
        'Comisión Proveedor',
        'Ganancia App',
        'Método Pago',
        'Fecha Creación',
        'Fecha Entrega',
        'Cancelado Por',
    ])

    # Optimizar queryset
    pedidos = queryset.select_related(
        'cliente__user',
        'proveedor',
        'repartidor__user'
    )

    # Datos
    for pedido in pedidos:
        writer.writerow([
            pedido.id,
            pedido.get_tipo_display(),
            pedido.get_estado_display(),
            pedido.cliente.user.get_full_name(),
            pedido.cliente.user.email,
            pedido.cliente.user.celular or 'N/A',
            pedido.proveedor.nombre if pedido.proveedor else 'N/A',
            pedido.repartidor.user.get_full_name() if pedido.repartidor else 'Sin asignar',
            pedido.direccion_entrega,
            f"{pedido.total:.2f}",
            f"{pedido.comision_repartidor:.2f}",
            f"{pedido.comision_proveedor:.2f}",
            f"{pedido.ganancia_app:.2f}",
            pedido.metodo_pago,
            pedido.creado_en.strftime('%Y-%m-%d %H:%M:%S'),
            pedido.fecha_entregado.strftime('%Y-%m-%d %H:%M:%S') if pedido.fecha_entregado else 'N/A',
            pedido.cancelado_por or 'N/A',
        ])

    logger.info(f"Reporte CSV generado: {len(pedidos)} pedidos")

    return response


# ============================================
# EXPORTAR ESTADÍSTICAS A EXCEL
# ============================================

def exportar_estadisticas_excel(data):
    """
    Exporta estadísticas resumidas a Excel

    Args:
        data: Dict con estadísticas

    Returns:
        HttpResponse con archivo Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Estadísticas"

    # Estilos
    header_font = Font(name='Arial', size=12, bold=True)
    data_font = Font(name='Arial', size=11)

    # Título
    ws['A1'] = 'ESTADÍSTICAS DE PEDIDOS'
    ws['A1'].font = Font(name='Arial', size=14, bold=True)

    # Datos
    row = 3
    for key, value in data.items():
        ws.cell(row=row, column=1).value = key.replace('_', ' ').title()
        ws.cell(row=row, column=1).font = header_font

        ws.cell(row=row, column=2).value = value
        ws.cell(row=row, column=2).font = data_font

        row += 1

    # Ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"estadisticas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


# ============================================
# CALCULAR RESUMEN FINANCIERO
# ============================================

def calcular_resumen_financiero(queryset):
    """
    Calcula resumen financiero de un queryset de pedidos

    Args:
        queryset: QuerySet de Pedido

    Returns:
        Dict con resumen financiero
    """
    from django.db.models import Sum, Avg, Count
    from pedidos.models import EstadoPedido

    # Solo pedidos entregados
    entregados = queryset.filter(estado=EstadoPedido.ENTREGADO)

    resumen = entregados.aggregate(
        total_ventas=Sum('total'),
        total_comision_repartidor=Sum('comision_repartidor'),
        total_comision_proveedor=Sum('comision_proveedor'),
        total_ganancia_app=Sum('ganancia_app'),
        ticket_promedio=Avg('total'),
        cantidad_pedidos=Count('id')
    )

    return {
        'total_ventas': resumen['total_ventas'] or 0,
        'total_comision_repartidor': resumen['total_comision_repartidor'] or 0,
        'total_comision_proveedor': resumen['total_comision_proveedor'] or 0,
        'total_ganancia_app': resumen['total_ganancia_app'] or 0,
        'ticket_promedio': round(resumen['ticket_promedio'] or 0, 2),
        'cantidad_pedidos': resumen['cantidad_pedidos'] or 0,
    }


# ============================================
# VALIDAR FECHAS DE REPORTE
# ============================================

def validar_rango_fechas(fecha_inicio, fecha_fin):
    """
    Valida que el rango de fechas sea correcto

    Args:
        fecha_inicio: Date
        fecha_fin: Date

    Returns:
        Tuple (bool, str): (es_valido, mensaje_error)
    """
    if not fecha_inicio or not fecha_fin:
        return True, None

    if fecha_inicio > fecha_fin:
        return False, "La fecha de inicio no puede ser posterior a la fecha fin"

    # Validar que no sea muy amplio (máx 1 año)
    diferencia = (fecha_fin - fecha_inicio).days
    if diferencia > 365:
        return False, "El rango de fechas no puede superar 1 año"

    return True, None
